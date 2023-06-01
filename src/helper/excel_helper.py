from openpyxl import Workbook
import openpyxl

INDEX_ROW = 0
DATE_ROW = 1
CATEGORY_ROW = 2
USE_ROW = 3
KEYWORD_ROW = 4
PRICE_ROW = 5

def build_excel_formula(col: str, row: str):
    return "=IFERROR(SUMIF(INDIRECT(\"\'\"&" + col + "$1&\" \"&$A$1&\"\'!\"&IF(ISBLANK($A" + row + "),\"E2:E\",\"C2:C\")&MAX(INDIRECT(\"\'\"&" + col + "$1&\" \"&$A$1&\"\'!\"&\"A:A\"))+1),IF(ISBLANK($A" + row + "),$B" + row + ",$A" + row + "),INDIRECT(\"\'\"&" + col + "$1&\" \"&$A$1&\"\'!\"&\"F2:F\"&MAX(INDIRECT(\"\'\"&" + col + "$1&\" \"&$A$1&\"\'!\"&\"A:A\"))+1)), 0)"
    #return "=IFERROR(SUMIF(INDIRECT(\"\'\"&" + col + "$1&\" \"&$A$1&\"\'!\"&IF(ISBLANK($"+ col + row + "),\"E2:E\",\"C2:C\")&MAX(INDIRECT(\"\'\"&" + col + "$1&\" \"&$A$1&\"\'!\"&\"A:A\"))+1),IF(ISBLANK($A8),$B8,$A8),INDIRECT(\"\'\"&C$1&\" \"&$A$1&\"\'!\"&\"F2:F\"&MAX(INDIRECT(\"\'\"&C$1&\" \"&$A$1&\"\'!\"&\"A:A\"))+1)), 0)"
 
def char_range(c1, c2):
    """Generates the characters from `c1` to `c2`, inclusive."""
    for c in range(ord(c1), ord(c2)+1):
        yield chr(c)

def load_categories_from_excel(excel_workbook: Workbook, all = False, with_index = False):
    categories = {}
    last_category = ''
    overview_sheet = excel_workbook['Overview'] # load worksheet overview
    
    category_area = False
    keywords = []
    index = 0
    index_category = 0
    for row in overview_sheet.iter_rows(max_col=2, values_only=True):
        index += 1
        
        if category_area == False and row[0] == None and row[1] == None: # a row with double None is the beginning of the categories
            category_area = True
            continue

        if category_area and row[0] != None:
            if last_category != '':
                if len(keywords) != 0 or all: # categories without any keywords are unnecessary to check
                    categories[last_category] = [index_category, keywords] if with_index else keywords

            last_category = row[0]
            index_category = index
            keywords = []
        
        if category_area and row[1] != None:
            keywords.append(row[1])

        if category_area and row[0] == None and row[1] == None: # a row with double None is also the end of the categories
            categories[last_category] = [index_category, keywords] if with_index else keywords
            break
    
    print(categories)
    return categories

def check_for_manual_changes(excel_file: Workbook, new_dealings, sheet_name):

    if sheet_name in excel_file.get_sheet_names(): # check if a sheet with the name already exists
        old_sheet = excel_file.get_sheet_by_name(sheet_name)

        i = 0
        for old_row in old_sheet.iter_rows(min_row=2,  values_only=True):
            if new_dealings[i][CATEGORY_ROW] == 'Sonstiges' and old_row[CATEGORY_ROW] != 'Sonstiges':
                new_dealings[i][CATEGORY_ROW] = old_row[CATEGORY_ROW]
                new_dealings[i][KEYWORD_ROW] = old_row[KEYWORD_ROW]

            i += 1
    
    return new_dealings


def export_to_excel(excel_file: Workbook, first_row, dealings, sheet_name, path_excel):

    if sheet_name in excel_file.get_sheet_names():
        sheet = excel_file.get_sheet_by_name(sheet_name)
        excel_file.remove_sheet(sheet)

    excel_file.create_sheet(sheet_name)
    sheet = excel_file.get_sheet_by_name(sheet_name)
    sheet.append(first_row)

    for row in dealings:
        sheet.append(row)

    excel_file.save(path_excel)
    
    
def insert_formulas(excel_sheet, index_str: str):
    for col in char_range('C', 'N'):
        position = col + index_str
        excel_sheet[position].value = build_excel_formula(col, index_str)
        
    excel_sheet["P" + index_str].value = "=SUM(C" + index_str + ":N" + index_str + ")"
    excel_sheet["Q" + index_str].value = "=SUM(C" + index_str + ":N" + index_str + ")/COUNTA(C" + index_str + ":N" + index_str + ")"
    
    excel_sheet["P" + index_str].number_format = "0.00"
    excel_sheet["Q" + index_str].number_format = "0.00"
    
def update_formulas(excel_sheet, start):
    index = start +1
    while not (excel_sheet["A" + str(index)].value == None and excel_sheet["B" + str(index)].value == None):
        
        insert_formulas(excel_sheet, str(index))
        index += 1
        
def get_index_of_category(excel_sheet, category) -> int or None:
    index = 1
    
    while excel_sheet['A' + str(index)].value != category:
        if excel_sheet['A' + str(index)].value == 'Sonstiges':
            index = None
            break
        index += 1
    
    return index
    

def export_new_categories(categories_new, path_excel):
    excel_workbook = openpyxl.load_workbook(path_excel)
    overview_sheet = excel_workbook['Overview']
    
    categories_excel = load_categories_from_excel(excel_workbook, True, True)
    
    for new in categories_new:
        if new in categories_excel.keys():
            print("The category already exists appending keywords")
            index = get_index_of_category(overview_sheet, new) + 1
            index_str = str(index)
            
            for keyword in categories_new[new]:
                overview_sheet.insert_rows(index)
                position = "B" + str(index_str)
                overview_sheet[position] = keyword
            
        else:
            print("New category")
            index = get_index_of_category(overview_sheet, 'Sonstiges') # the index is changing when something gets inserted
            index_str = str(index)
            
            overview_sheet.insert_rows(index)
            position = "A" + index_str
            overview_sheet[position] = new
            
            index += 1
            index_str = str(index)
            if len(categories_new[new]) > 0:
                for keyword in categories_new[new]:
                    overview_sheet.insert_rows(index)
                    position = "B" + str(index_str)
                    overview_sheet[position] = keyword
    
    update_formulas(overview_sheet, 7)
    excel_workbook.save(path_excel)

    


    