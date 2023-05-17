import unittest
import openpyxl

import src.helper.config_helper as config_helper
import src.helper.excel_helper as excel_helper

class ExampleTest(unittest.TestCase):

    def test_example(self):
        self.assertEqual(1,1, "Should be 1")
        
class ConfigHelper(unittest.TestCase):
    
    def test_search_for_new_categories(self):
        dealings = []
        
        # Test if new categories are found
        dealings.append([0, '3.04.2023', 'Sonstiges', 'Rewe', '', '30'])
        dealings.append([0, '3.04.2023', 'Sonstiges', 'Netflix', '', '30'])
        dealings.append([0, '3.04.2023', 'Sonstiges', 'Eventim', '', '30'])
        dealings.append([0, '3.04.2023', 'Sonstiges', 'Das ist ein Test', '', '30'])
        
        found_categories = config_helper.search_for_new_categories(dealings)
        self.assertEqual({'Lebensmittel': ['Rewe'], 'Streaming': ['Netflix'], 'Freizeit': ['Eventim']}, found_categories)
        
        
        # test if the appending of new categories is working
        dealings2 = []
        dealings2.append([0, '3.04.2023', 'Sonstiges', 'Das ist ein Test', '', '30'])
        dealings2.append([0, '3.04.2023', 'Sonstiges', 'Tegut', '', '30'])
        dealings2.append([0, '3.04.2023', 'Sonstiges', 'WOW', '', '30'])
        dealings2.append([0, '3.04.2023', 'Sonstiges', 'Das ist ein weiterer Test', '', '30'])
        
        found_categories = config_helper.search_for_new_categories(dealings2, found_categories)
        print(found_categories)
        self.assertEqual({'Lebensmittel': ['Rewe', 'Tegut'], 'Streaming': ['Netflix', 'WOW'], 'Freizeit': ['Eventim']}, found_categories)
        
        
        # test if the return value is correct if there are not new categories
        dealings[0][2] = 'Lebensmittel'
        dealings[1][2] = 'Streaming'
        dealings[2][2] = 'Freizeit'
        
        found_categories = config_helper.search_for_new_categories(dealings, None)
        self.assertEqual(None, found_categories)
        
        

class ExcelTest(unittest.TestCase):
    
    def test_get_index_of_category(self):
        # test if the get index works properly
        workbook = openpyxl.Workbook()
        workbook.create_sheet('Test')
        sheet = workbook['Test']
        
        sheet['A1'] = 'Essen'
        sheet['A4'] = 'Lebensmittel'
        sheet['A7'] = 'Sonstiges'
        
        self.assertEqual(excel_helper.get_index_of_category(sheet, 'Essen'), 1)
        self.assertEqual(excel_helper.get_index_of_category(sheet, 'Not in Excel'), None)
        self.assertEqual(excel_helper.get_index_of_category(sheet, 'Sonstiges'), 7)
        
        
    
    def test_new_categories(self):
        path_excel = config_helper.get_excel_path()
        
        categories_new = {"Test": ["Keyword_1", "Test2"], "Lebensmittel": ["Test"]}
        
        excel_helper.export_new_categories(categories_new, path_excel)
        
        
""" class ExportCategories(unittest.TestCase):
    
    def test_export_config(self):
        path_excel = config_helper.get_excel_path()
        
        workbook = openpyxl.load_workbook(path_excel)
        categories = excel_helper.load_categories_from_excel(workbook, all=True)
        config_helper.safe_categories(categories) """
        
        
        
        
        


if __name__ == '__main__':
    unittest.main()