import unittest
import openpyxl

import config_helper
import excel_helper

class ExampleTest(unittest.TestCase):

    def test_example(self):
        self.assertEqual(1,1, "Should be 1")

class ExcelTest(unittest.TestCase):
    
    def test_get_index_of_category(self):
        workbook = openpyxl.Workbook()
        workbook.create_sheet('Test')
        sheet = workbook['Test']
        
        sheet['A1'] = 'Essen'
        sheet['A4'] = 'Lebensmittel'
        sheet['A7'] = 'Sonstiges'
        
        self.assertEqual(excel_helper.get_index_of_category(sheet, 'Essen'), 1)
        self.assertEqual(excel_helper.get_index_of_category(sheet, 'Not in Excel'), None)
        self.assertEqual(excel_helper.get_index_of_category(sheet, 'Sonstiges'), 7)
        
        
    
    def test_new_categories(self):
        path_excel = config_helper.get_excel_path()
        
        categories_new = {"Test": ["Keyword_1", "Test2"], "Lebensmittel": ["Test"]}
        
        excel_helper.export_new_categories(categories_new, path_excel)
        
        
""" class ExportCategories(unittest.TestCase):
    
    def test_export_config(self):
        path_excel = config_helper.get_excel_path()
        
        workbook = openpyxl.load_workbook(path_excel)
        categories = excel_helper.load_categories_from_excel(workbook, all=True)
        config_helper.safe_categories(categories) """
        
        
        
        
        


if __name__ == '__main__':
    unittest.main()