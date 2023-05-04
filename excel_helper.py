from openpyxl import Workbook

INDEX_ROW = 0
DATE_ROW = 1
CATEGORY_ROW = 2
USE_ROW = 3
KEYWORD_ROW = 4
PRICE_ROW = 5

def load_categories_from_excel(excel_workbook: Workbook):
    categories = {}
    last_category = ''
    overview_sheet = excel_workbook.get_sheet_by_name('Overview') # load worksheet overview
    
    category_area = False
    keywords = []
    for row in overview_sheet.iter_rows(max_col=2, values_only=True):

        if category_area == False and row[0] == None and row[1] == None: # a row with double None is the beginning of the categories
            category_area = True
            continue

        if category_area and row[0] != None:
            if len(keywords) != 0: # categories without any keywords are unnecessary to check
                categories[last_category] = keywords

            last_category = row[0]
            keywords = []
        
        if category_area and row[1] != None:
            keywords.append(row[1])

        if category_area and row[0] == None and row[1] == None: # a row with double None is also the end of the categories
            break
    
    print(categories)
    return categories

def check_for_manual_changes(excel_file: Workbook, new_dealings, sheet_name):

    if sheet_name in excel_file.get_sheet_names(): # check if a sheet with the name already exists
        old_sheet = excel_file.get_sheet_by_name(sheet_name)

        i = 0
        for old_row in old_sheet.iter_rows(min_row=2,  values_only=True):
            if new_dealings[i][CATEGORY_ROW] == 'Sonstiges' and old_row[CATEGORY_ROW] != 'Sonstiges':
                new_dealings[i][CATEGORY_ROW] = old_row[CATEGORY_ROW]
                new_dealings[i][KEYWORD_ROW] = old_row[KEYWORD_ROW]

            i += 1
    
    return new_dealings


def export_to_excel(excel_file: Workbook, first_row, dealings, sheet_name, path_excel):

    if sheet_name in excel_file.get_sheet_names():
        sheet = excel_file.get_sheet_by_name(sheet_name)
        excel_file.remove_sheet(sheet)

    excel_file.create_sheet(sheet_name)
    sheet = excel_file.get_sheet_by_name(sheet_name)
    sheet.append(first_row)

    for row in dealings:
        sheet.append(row)

    excel_file.save(path_excel)

def export_new_categories(categories, path_excel):
        pass 

    


    